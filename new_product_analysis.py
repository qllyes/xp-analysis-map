import streamlit as st
import pandas as pd
import json
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Optional, Dict, Any, Tuple
import logging
import warnings
import re
import pymysql
from sqlalchemy import create_engine

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.filterwarnings('ignore', category=UserWarning, module='pandas.io.sql')

# --- 常量定义 (Constants) ---
CONFIG_FILE = Path("config.json")

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# --- 文件处理类 (File Processing) ---
class FileProcessor:
    """文件处理类，负责Excel文件的读取和保存"""
    
    @staticmethod
    def read_excel_safe(file_path_or_buffer) -> pd.DataFrame:
        """
        安全的Excel读取方法，先将文件保存为临时文件，然后从临时文件读取
        """
        import io
        import tempfile
        import os

        # 获取文件内容到BytesIO
        if hasattr(file_path_or_buffer, 'getvalue'):
            original_buffer = io.BytesIO(file_path_or_buffer.getvalue())
            file_name = file_path_or_buffer.name
        else:
            with open(file_path_or_buffer, 'rb') as f:
                original_buffer = io.BytesIO(f.read())
            file_name = str(file_path_or_buffer)

        # 创建临时文件
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            temp_file_path = temp_file.name
            original_buffer.seek(0)
            temp_file.write(original_buffer.read())

        try:
            # 尝试使用不同引擎读取临时文件
            engines = ['openpyxl', 'xlrd']
            for engine in engines:
                try:
                    return pd.read_excel(temp_file_path, engine=engine)
                except Exception as e:
                    logger.warning(f"读取失败 (引擎 {engine}): {str(e)}")
                    continue

            raise ValueError(f"无法读取文件 {file_name}。请确认文件格式正确。")
            
        finally:
            # 清理临时文件
            try:
                os.unlink(temp_file_path)
            except Exception as cleanup_e:
                logger.warning(f"临时文件清理失败: {str(cleanup_e)}")

# --- 文件上传组件类 (File Upload Widget) ---
class FileUploadWidget:
    """文件上传组件类，提供更好的状态管理"""
    
    def __init__(self, config_manager=None):
        # 为了保持兼容性，即使没有配置管理器也能工作
        self.config_manager = config_manager
    
    def render(self, label: str, session_state_key: str) -> Optional[pd.DataFrame]:
        """
        渲染文件上传组件并返回加载的数据
        
        Args:
            label: 上传组件的标签
            session_state_key: session state中保存数据的键
            
        Returns:
            加载的DataFrame或None
        """
        # 初始化session state
        if session_state_key not in st.session_state:
            st.session_state[session_state_key] = None
        
        # 创建唯一的uploader key
        uploader_key = f"uploader_{session_state_key}_{hash(label)}"
        
        # 跟踪文件状态变化
        file_state_key = f"file_state_{session_state_key}"
        if file_state_key not in st.session_state:
            st.session_state[file_state_key] = None
        
        # 文件上传器
        uploaded_file = st.file_uploader(
            label, 
            type=['xlsx', 'xls'], 
            key=uploader_key,
            help="支持.xlsx和.xls格式的Excel文件"
        )
        
        # 处理文件上传或删除
        if uploaded_file is not None:
            # 文件被上传
            st.session_state[file_state_key] = uploaded_file.name
            return self._handle_file_upload(uploaded_file, session_state_key)
        else:
            # 检查是否文件被删除
            if st.session_state.get(file_state_key) is not None and st.session_state.get(session_state_key) is not None:
                st.session_state[session_state_key] = None
                st.session_state[file_state_key] = None
        
        # 添加手动文件路径输入作为备选方案
        with st.expander("🔧 高级选项"):
            manual_path = st.text_input(
                "手动指定文件路径:",
                key=f"manual_path_{session_state_key}",
                help="例如: C:/Users/用户名/Desktop/文件.xlsx"
            )
            
            if st.button("📁 从路径加载", key=f"load_manual_{session_state_key}"):
                if manual_path and Path(manual_path).exists():
                    try:
                        with st.spinner("正在读取文件..."):
                            data = FileProcessor.read_excel_safe(Path(manual_path))
                        
                        st.session_state[session_state_key] = data
                        st.session_state[file_state_key] = Path(manual_path).name
                        
                        st.success(f"✅ 成功加载: {Path(manual_path).name} ({len(data)} 行)")
                        return data
                        
                    except Exception as e:
                        st.error(f"❌ 读取失败: {str(e)}")
                        return None
                elif manual_path:
                    st.error("❌ 文件路径不存在")
                else:
                    st.warning("⚠️ 请输入有效路径")
        
        return st.session_state.get(session_state_key)
    
    def _handle_file_upload(self, uploaded_file, session_state_key: str) -> pd.DataFrame:
        """处理文件上传"""
        try:
            # 检查文件大小
            file_size = len(uploaded_file.getvalue())
            if file_size > 50 * 1024 * 1024:  # 50MB限制
                st.warning("⚠️ 文件较大，读取可能需要一些时间...")
            
            # 使用临时文件方式读取
            with st.spinner(f"正在读取文件 {uploaded_file.name}..."):
                data = FileProcessor.read_excel_safe(uploaded_file)
            
            # 更新session state
            st.session_state[session_state_key] = data
            
            st.success(f"✅ 成功加载文件: {uploaded_file.name} ({len(data)} 行, {len(data.columns)} 列)")
            return data
            
        except Exception as e:
            st.error(f"❌ 读取失败: {uploaded_file.name}")
            st.info("💡 建议：检查文件格式或使用高级选项")
            
            st.session_state[session_state_key] = None
            return None

# --- SQL处理器类 ---
class SQLProcessor:
    """SQL处理器类，负责执行SQL查询"""
    
    @staticmethod
    def read_sql_file(file_path: Path) -> str:
        """读取SQL文件内容"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        except Exception as e:
            logger.error(f"读取SQL文件失败: {e}")
            raise ValueError(f"读取SQL文件失败: {str(e)}")
    
    @staticmethod
    def extract_sql_from_file(file_path: Path) -> str:
        """从SQL文件中提取查询语句"""
        sql_content = SQLProcessor.read_sql_file(file_path)
        # 移除注释
        sql_content = re.sub(r'--.*$', '', sql_content, flags=re.MULTILINE)
        sql_content = re.sub(r'/\*.*?\*/', '', sql_content, flags=re.DOTALL)
        return sql_content.strip()
    
    @staticmethod
    def execute_sql_query(sql_query: str) -> pd.DataFrame:
        """执行SQL查询并返回结果（连接MySQL数据库）"""
        # 注意：这里使用了需求文档中提到的默认SQL文件
        default_sql_file = Path("对标品.sql")
        if default_sql_file.exists():
            try:
                with open(default_sql_file, 'r', encoding='utf-8') as f:
                    sql_query = f.read()
            except Exception as e:
                logger.error(f"读取默认SQL文件失败: {e}")
                st.error(f"读取默认SQL文件失败: {str(e)}")
                return pd.DataFrame()
        
        host = "10.243.0.221"      # 数据库地址
        port = 3306                # 端口号
        user = "xinpin"     # 用户名
        password = "xinpin" # 密码
        database = "new_goods_manage" # 数据库名

        try:
            # 建立数据库连接
            conn = pymysql.connect(
                host=host,
                port=port,
                user=user,
                password=password,
                database=database,
                charset='utf8mb4'
            )
            # 执行SQL查询
            engine = create_engine(
                "mysql+pymysql://xinpin:xinpin@10.243.0.221:3306/new_goods_manage?charset=utf8mb4"
            )
            df = pd.read_sql(sql_query, engine)
            conn.close()
            return df
        except Exception as e:
            logger.error(f"执行SQL查询失败: {e}")
            st.error(f"执行SQL查询失败: {str(e)}")
            return pd.DataFrame()

# --- 映射处理类 (Mapping Processor) ---
class MappingProcessor:
    """映射处理类，负责字段映射逻辑"""
    
    @staticmethod
    def run_mapping(map_df: pd.DataFrame, source_df: pd.DataFrame, source_type: str = 'table2') -> pd.DataFrame:
        """执行字段映射
        Args:
            map_df: 映射关系表（包含三列：目标字段名、table2字段名、table3字段名）
            source_df: 源数据表
            source_type: 源数据类型 ('table2' 或 'table3')
        """
        if map_df.empty or source_df.empty:
            raise ValueError("映射表或源数据为空")
        
        # 提取映射关系
        # 根据source_type选择对应的映射列
        if source_type == 'table2':
            # table2字段名在第二列（索引1）
            source_field_col = 1
        elif source_type == 'table3':
            # table3字段名在第三列（索引2）
            source_field_col = 2
        else:
            raise ValueError("source_type 必须是 'table2' 或 'table3'")
        
        map_df_clean = map_df.dropna(how='all') # 移除 NaN
        # 目标字段名在第一列（索引0）
        main_fields = map_df_clean.iloc[:, 0].tolist()
        # 根据source_type选择源字段列
        source_fields = map_df_clean.iloc[:, source_field_col].tolist()
        
        
        # 构建映射字典 {源字段: 目标字段}
        field_map = {}
        for target_field, source_field in zip(main_fields, source_fields):
            # 只处理非空且有效的映射
            if pd.notna(source_field) and source_field != '' and pd.notna(target_field):
                field_map[str(source_field)] = str(target_field)
        
        
        # 从 source_df 中选取并重命名存在的列
        available_cols = [col for col in field_map if col in source_df.columns]
        result_df = source_df[available_cols].rename(columns=field_map)
        
        # 添加缺失的目标列，并填充为空字符串（仅添加在映射清单中出现过的目标字段）
        existing_cols = set(result_df.columns)
        for col in main_fields:
            if col not in existing_cols:
                result_df[col] = ''
        
        # 重新排序列以匹配目标字段顺序
        result_df = result_df[main_fields]
        
        return result_df

# --- 数据筛选处理器 ---
class DataFilterProcessor:
    """数据筛选处理器，负责对标品数据的筛选"""
    
    @staticmethod
    def filter_benchmark_data(benchmark_df: pd.DataFrame, scm_df: pd.DataFrame) -> pd.DataFrame:
        """根据SCM数据筛选对标品数据"""
        if benchmark_df.empty or scm_df.empty:
            raise ValueError("对标品数据或SCM数据为空")
        
        # 获取SCM数据中的通用名和三级大类
        # 根据需求文档，SCM数据中对应的列为"通用名"和"策略分类"
        scm_common_names = set(scm_df.get('通用名', pd.Series(dtype=object)).dropna().astype(str).unique())
        scm_category3_names = set(scm_df.get('策略分类', pd.Series(dtype=object)).dropna().astype(str).unique())
        
        # 筛选对标品数据
        # 对标品数据中对应的列为"通用名"和"三级策略分类"
        filtered_df = benchmark_df[
            (benchmark_df['通用名'].astype(str).isin(scm_common_names)) |
            (benchmark_df['三级策略分类'].astype(str).isin(scm_category3_names))
        ]
        
        return filtered_df

# --- 数据合并处理器 ---
class DataMerger:
    """数据合并处理器，负责合并映射后的数据并排序"""
    
    @staticmethod
    def merge_and_sort_data(map_scm_df: pd.DataFrame, map_benchmark_df: pd.DataFrame) -> pd.DataFrame:
        """合并映射后的数据并按指定规则排序"""
        # 创建副本以避免修改原始数据
        map_scm_df = map_scm_df.copy()
        map_benchmark_df = map_benchmark_df.copy()
        
        if map_scm_df.empty and map_benchmark_df.empty:
            raise ValueError("两个映射后的数据表都为空")
        
        # 合并数据
        if not map_scm_df.empty and not map_benchmark_df.empty:
            # 添加标识列区分数据来源
            map_scm_df['__source__'] = 'scm'
            map_benchmark_df['__source__'] = 'benchmark'
            
            # 合并数据
            merged_df = pd.concat([map_scm_df, map_benchmark_df], ignore_index=True)
            
            # 按三级大类分组排序
            if '三级大类' in merged_df.columns:
                # 对benchmark数据按销售额降序排序
                benchmark_part = merged_df[merged_df['__source__'] == 'benchmark'].copy()
                if '近90天月均销售金额' in benchmark_part.columns:
                    benchmark_part = benchmark_part.sort_values('近90天月均销售金额', ascending=False)
                
                # SCM数据在前，benchmark数据在后
                scm_part = merged_df[merged_df['__source__'] == 'scm']
                
                # 重新组合数据
                result_parts = []
                # 按三级大类分组处理
                for category in merged_df['三级大类'].dropna().unique():
                    scm_category = scm_part[scm_part['三级大类'] == category]
                    benchmark_category = benchmark_part[benchmark_part['三级大类'] == category]
                    if not scm_category.empty or not benchmark_category.empty:
                        result_parts.append(pd.concat([scm_category, benchmark_category], ignore_index=True))
                
                # 处理没有三级大类的数据
                scm_no_category = scm_part[scm_part['三级大类'].isna()]
                benchmark_no_category = benchmark_part[benchmark_part['三级大类'].isna()]
                if not scm_no_category.empty or not benchmark_no_category.empty:
                    result_parts.append(pd.concat([scm_no_category, benchmark_no_category], ignore_index=True))
                
                # 合并所有部分
                if result_parts:
                    final_df = pd.concat(result_parts, ignore_index=True)
                else:
                    final_df = merged_df
            else:
                # 如果没有三级大类字段，直接合并
                final_df = merged_df
            
            # 删除标识列
            final_df = final_df.drop(columns=['__source__'])
        elif not map_scm_df.empty:
            final_df = map_scm_df
        else:
            final_df = map_benchmark_df
        
        return final_df

# --- 结果导出类 (Result Exporter) ---
class ResultExporter:
    """结果导出类，负责生成和下载结果文件"""
    
    @staticmethod
    def export_to_excel(df: pd.DataFrame, scm_indices: list) -> Tuple[BytesIO, str]:
        """导出DataFrame到Excel格式，并对SCM数据行添加黄色背景"""
        output = BytesIO()
        
        # 创建Excel写入器
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='目标表')
            
            # 获取工作表
            worksheet = writer.sheets['目标表']
            
            # 设置黄色背景样式
            from openpyxl.styles import PatternFill
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            
            # 对SCM数据行应用黄色背景（注意：Excel行索引从1开始，且有标题行）
            for row_idx in scm_indices:
                for col_idx in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row_idx + 2, column=col_idx).fill = yellow_fill
        
        filename = f'目标表_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        
        return output, filename

# --- 主应用类 (Main Application) ---
class NewProductAnalysisApp:
    """新品分析应用主类"""
    
    def __init__(self):
        self.file_processor = FileProcessor()
        self.upload_widget = FileUploadWidget()
        self.mapping_processor = MappingProcessor()
        self.sql_processor = SQLProcessor()
        self.filter_processor = DataFilterProcessor()
        self.data_merger = DataMerger()
        self.result_exporter = ResultExporter()
    
    def _inject_custom_css(self):
        """注入自定义CSS，实现卡片式布局和美化"""
        st.markdown(
            """
            <style>
                .card { background-color: #f9fafb; border-radius: 12px; padding: 20px; margin-bottom: 18px; box-shadow: 0 2px 8px rgba(0,0,0,0.06); }
                .card-title { font-size: 16px; font-weight: 700; color: #333; margin-bottom: 12px; }
                .muted { color: #666; font-size: 13px; }
                .primary-btn .st-emotion-cache-7ym5gk { font-size: 16px; padding: 10px 16px; }
            </style>
            """,
            unsafe_allow_html=True,
        )
    
    def render_header(self):
        """渲染页面头部"""
        st.set_page_config(page_title="新品过会分析表生成工具", layout="wide")
        st.title("🔍 新品过会分析表生成工具")
        st.markdown(
            """
            **操作步骤：** 
            1. 上传映射关系表 → 
            2. 上传新品申报数据 → 
            3. 运行生成 → 
            4. 下载结果
            """
        )
        st.divider()
    
    def render_input_section(self):
        """① 数据上传（两列并排，移除SQL可视项）"""
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<div class="card-title">① 数据上传</div>', unsafe_allow_html=True)
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**映射关系表**")
            map_file = st.file_uploader("上传定义字段映射关系的Excel文件", type=["xlsx", "xls"], key="map_uploader_new")
            if map_file:
                with st.spinner("正在读取映射表..."):
                    st.session_state["map_df"] = FileProcessor.read_excel_safe(map_file)
        with col2:
            st.markdown("**新品申报数据**")
            scm_file = st.file_uploader("上传从SCM系统导出的新品申报Excel文件", type=["xlsx", "xls"], key="scm_uploader_new")
            if scm_file:
                with st.spinner("正在读取新品数据..."):
                    st.session_state["scm_df"] = FileProcessor.read_excel_safe(scm_file)
        # 成功提示（行+列）
        if st.session_state.get("map_df") is not None:
            df = st.session_state["map_df"]
            st.success(f"✅ 映射关系表已加载 ({df.shape[0]} 行，{df.shape[1]} 列)")
        if st.session_state.get("scm_df") is not None:
            df = st.session_state["scm_df"]
            st.success(f"✅ 新品申报数据已加载 ({df.shape[0]} 行，{df.shape[1]} 列)")
        st.markdown('</div>', unsafe_allow_html=True)
    
    def render_action_section(self):
        """② 执行分析（单独卡片 + 大按钮）"""
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<div class="card-title">② 执行生成</div>', unsafe_allow_html=True)
        if st.button("🚀 运行", type="primary", use_container_width=True):
            self._process_analysis()
        st.markdown('</div>', unsafe_allow_html=True)
    
    def _process_analysis(self):
        """核心处理流程（SQL后台静默加载，仅在失败时提示）"""
        map_df = st.session_state.get("map_df")
        scm_df = st.session_state.get("scm_df")
        if map_df is None or scm_df is None:
            st.error("❌ 请先上传映射关系表和新品申报数据！")
            return
        try:
            status = st.status("准备开始生成…", expanded=True)
            # 步骤1：查询对标品数据
            status.update(label="🔎 正在从数据库查询对标品数据…", state="running")
            sql_path = Path("对标品.sql")
            if not sql_path.exists():
                status.update(label="❌ 后台SQL文件缺失，请联系维护人员。", state="running")
                st.error("❌ 后台SQL文件缺失，请联系维护人员。")
                return
            sql_query = self.sql_processor.read_sql_file(sql_path)
            benchmark_df = self.sql_processor.execute_sql_query(sql_query)
            if benchmark_df.empty:
                st.warning("⚠️ 对标品数据查询为空，结果将仅包含新品数据。")
            
            # 步骤2：映射 + 筛选 + 合并
            status.update(label="🧭 正在进行映射转换与数据合并…", state="running")
            map_scm_df = self.mapping_processor.run_mapping(map_df.copy(), scm_df.copy(), source_type='table2')
            filtered_benchmark_df = (
                self.filter_processor.filter_benchmark_data(benchmark_df.copy(), scm_df.copy())
                if not benchmark_df.empty else pd.DataFrame()
            )
            map_benchmark_df = (
                self.mapping_processor.run_mapping(map_df.copy(), filtered_benchmark_df.copy(), source_type='table3')
                if not filtered_benchmark_df.empty else pd.DataFrame()
            )
            target_df = self.data_merger.merge_and_sort_data(map_scm_df.copy(), map_benchmark_df.copy())
            
            # 步骤3：导出并完成
            status.update(label="📦 正在生成Excel文件…", state="running")
            scm_indices = list(range(len(map_scm_df)))
            output, filename = self.result_exporter.export_to_excel(target_df, scm_indices)
            st.session_state["result_df"] = target_df
            st.session_state["result_output"] = output
            st.session_state["result_filename"] = filename
            status.update(label="🎉 生成完成！", state="complete")
            #st.success("🎉 生成完成！")
        except Exception as e:
            st.error(f"❌ 分析过程中发生错误: {e}")
            logger.error(f"分析处理错误: {e}", exc_info=True)
    2
    def render_results_section(self):
        """③ 分析结果（仅在有结果时显示）"""
        if st.session_state.get("result_df") is None:
            return
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<div class="card-title">③ 生成结果</div>', unsafe_allow_html=True)
        st.success("生成成功，以下是结果摘要与下载：")
        col1, col2 = st.columns(2)
        with col1:
            st.metric("总计行数", st.session_state["result_df"].shape[0])
        with col2:
            st.metric("总计列数", st.session_state["result_df"].shape[1])
        st.download_button(
            label="📥 下载Excel结果文件",
            data=st.session_state["result_output"].getvalue(),
            file_name=st.session_state["result_filename"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        with st.expander("点击预览结果数据"):
            st.dataframe(st.session_state["result_df"])
        st.markdown('</div>', unsafe_allow_html=True)
    
    def run(self):
        """运行应用"""
        self.render_header()
        self._inject_custom_css()
        self.render_input_section()
        self.render_action_section()
        self.render_results_section()

# --- 主函数 ---
def main():
    """应用入口函数"""
    app = NewProductAnalysisApp()
    app.run()

if __name__ == "__main__":
    main()