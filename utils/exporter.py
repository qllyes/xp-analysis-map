import pandas as pd
from io import BytesIO
from datetime import datetime
from typing import Tuple, List
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.cell import MergedCell
from openpyxl.utils import get_column_letter

class ResultExporter:
    """结果导出类，负责生成和下载结果文件，并应用复杂的格式。"""

    @staticmethod
    def export_to_excel(df: pd.DataFrame, separator_indices: List[int], scm_indices: List[int], purchase_mode: str) -> Tuple[BytesIO, str]:
        """
        导出DataFrame到Excel，并应用所有指定的格式。
        该函数现在能正确处理 separator_indices 为空列表的情况（统采模式）。
        """
        output = BytesIO()
        
        # 统采模式下，不需要分隔符占位符，可以直接写入
        df_to_write = df.copy()
        if "_SEPARATOR_" in df_to_write.iloc[:, 0].values:
             df_to_write = df.replace("_SEPARATOR_", "")

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_to_write.to_excel(writer, index=False, sheet_name='目标表')
            
            workbook = writer.book
            worksheet = writer.sheets['目标表']

            # --- 定义样式 ---
            default_font = Font(name='微软雅黑', size=9)
            header_font = Font(name='微软雅黑', size=9, bold=True)
            red_font = Font(name='微软雅黑', size=9, color="FF0000")
            
            wrap_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True) 
            no_wrap_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            
            thin_border_side = Side(style='thin', color='000000')
            thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

            # --- 定义数字格式 ---
            percent_format = '0.00%;-0.00%;0.00%;@'
            decimal_2_format = '0.00;-0.00;0.00;@'
            decimal_1_format = '0.0;-0.0;0.0;@'
            integer_format = '0;-0;0;@'
            text_format = '@'
            
            column_formats = {
                '返利率(%)': percent_format, '通用名补偿后毛利率': percent_format,
                '日服/使用成交价（顾客）': decimal_2_format, '日服/使用底价': decimal_2_format,
                '标准单位进价': decimal_2_format, '标准单位底价': decimal_2_format,
                '标准单位成交价': decimal_2_format, '标准单位综合毛利额': decimal_2_format,
                '标准单位零售定价': decimal_2_format, '进价': decimal_2_format,
                '新品底价/对标品最低底价': decimal_2_format, '底价 *(返利后)': decimal_2_format,
                '近90天门店最新一批的底价': decimal_2_format,'近90天门店级最低底价': decimal_2_format, 
                '9000的移动平均价':decimal_2_format, '9000的最后进价':decimal_2_format,
                '【使用最新】近90天购进批次的最新底价-不含销售返利':decimal_2_format,
                '【使用最低】近90天购进批次的最低底价':decimal_2_format,
                '预估/实际成交价': decimal_1_format, '建议零售价': decimal_1_format,
                '过会编码': text_format, '新品编码': text_format, '国际条码': text_format,
                '近90天月均销售数量': integer_format, '近90天月均销售金额': integer_format,
                '近90天月均前台含税毛利额': integer_format, '近90天月均补偿后含税毛利额': integer_format,
                '超级旗舰店铺货商品数量': integer_format, '旗舰店铺货商品数量': integer_format,
                '大店铺货商品数量': integer_format, '中店铺货商品数量': integer_format,
                '小店铺货商品数量': integer_format, '成长店铺货商品数量': integer_format,
                '通用名月均销量': integer_format, '通用名月均销售额': integer_format,
                '通用名月均前台毛利额': integer_format, '通用名月均补偿后毛利额': integer_format,
            }
            red_font_columns = ['新品底价/对标品最低底价', '底价 *(返利后)']

            # --- 应用常规样式和格式 ---
            for col_idx, col_name in enumerate(df_to_write.columns, 1):
                header_cell = worksheet.cell(row=1, column=col_idx)
                header_cell.font = header_font
                header_cell.alignment = wrap_alignment
                
                for row_idx in range(2, len(df_to_write) + 2):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.alignment = no_wrap_alignment
                    cell.font = red_font if col_name in red_font_columns else default_font
                    if col_name in column_formats:
                        cell.number_format = column_formats[col_name]

            # --- 处理分隔行：背景色、合并、公式 (仅在地采模式下执行) ---
            if separator_indices:
                for sep_idx in separator_indices:
                    excel_row = sep_idx + 2
                    worksheet.row_dimensions[excel_row].height = 150
                    
                    # 地采逻辑：合并Y-AL列和AV-BF列
                    worksheet.merge_cells(start_row=excel_row, start_column=25, end_row=excel_row, end_column=38) # Y to AL
                    worksheet.merge_cells(start_row=excel_row, start_column=48, end_row=excel_row, end_column=58) # AV to BF
                    
                    worksheet.cell(row=excel_row, column=25).alignment = wrap_alignment
                    worksheet.cell(row=excel_row, column=48).alignment = wrap_alignment
                    
                    scm_data_row = excel_row + 1
                    
                    # 地采模式的新公式
                    formula1 = f'=I{scm_data_row}&CHAR(10)&"1.顾客：；"&CHAR(10)&"2.公司：；"&CHAR(10)&"3.市场分析：；"&CHAR(10)&"4.供应商条件："&DB{scm_data_row}&"，"&DE{scm_data_row}&"，"&DI{scm_data_row}&"；"&CHAR(10)&"5.医保："&CK{scm_data_row}&"，"&"支付价"&"："&CL{scm_data_row}&"；"&CHAR(10)&"6.铺货通道："&CV{scm_data_row}&"；"&CHAR(10)&"挑战点：1.；"&CHAR(10)&"修改点：1.；"'
                    formula2 = f'="【引进理由】"&L{scm_data_row}&CHAR(10)&"【成份】"&EX{scm_data_row}&CHAR(10)&"【适应症】"&EZ{scm_data_row}&CHAR(10)&"【卖点】"&FB{scm_data_row}&CHAR(10)&"【关键搜索词】"&FC{scm_data_row}'
                    
                    worksheet.cell(row=excel_row, column=25).value = formula1
                    worksheet.cell(row=excel_row, column=48).value = formula2
                    
                    # 在分隔行的A/C/O列添加公式
                    worksheet.cell(row=excel_row, column=1).value = f'="压测"&M{scm_data_row}'
                    worksheet.cell(row=excel_row, column=3).value = f'=C{scm_data_row}'
                    worksheet.cell(row=excel_row, column=15).value = f'=O{scm_data_row}'

            # --- 处理SCM行：背景色 (所有模式都需要) ---
            for scm_idx in scm_indices:
                excel_row = scm_idx + 2
                for i in range(1, len(df_to_write.columns) + 1):
                    worksheet.cell(row=excel_row, column=i).fill = yellow_fill
                
            # --- 后处理步骤 ---
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue
                    if cell.value is None:
                        cell.value = '-'

            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border

        output.seek(0)
        output_mode = '地采' if purchase_mode != '统采' else '统采'

        filename = f'{output_mode}新品过会分析表_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return output, filename
